from openpyxl import load_workbook
import time

# 1.打开 Excel 表格并获取表格名称
workbook = load_workbook(filename="wrfile\mydata_m.xlsx")
print(workbook.sheetnames)

# 2.通过 sheet 名称获取表格
sheet = workbook["Sheet1"]
print(sheet)

## 3.获取表格的尺寸大小(几行几列数据) 这里所说的尺寸大小，指的是 excel 表格中的数据有几行几列，针对的是不同的 sheet 而言。
# print(sheet.dimensions)

# # 4.获取表格内某个格子的数据
# # 1 sheet["A1"]方式
# cell1 = sheet["B1"]
# cell2 = sheet["C1"]
# print(cell1.value, cell2.value)

# # workbook.active 打开激活的表格; sheet["A1"] 获取 A1 格子的数据; cell.value 获取格子中的值;
# print(sheet['B2'].value,sheet['B6'].value)
# print(sheet['A'])

# # 4.2sheet.cell(row=, column=)方式
# cell1 = sheet.cell(row = 1,column = 2)
# cell2 = sheet.cell(row = 1,column = 3)
# print(cell1.value, cell2.value)

# # 5. 获取一系列格子
# # 获取 A1:C2 区域的值
# cell = sheet["A1:C6"]
# print(cell)
# for i in cell:
#    for j in i:
#        print(j.value)



t1 = time.time()
for i in sheet.iter_rows(min_row=1, max_row=32, min_col=1, max_col=13):
   print('----------------------')
   for j in i:
       print(j.value)
t2=time.time()
print('\n')
print("使用openpyxl工具包遍历数据耗时：%.2f 秒"%(t2-t1))
